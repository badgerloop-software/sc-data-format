"""Parse Firmware Signal Spreadsheet tabs into normalized signal records."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any

import openpyxl


@dataclass
class ParsedSignal:
    key: str
    sheet: str
    row: int
    data_type: str
    can_id: str
    bit_offset: int
    category: str
    units: str = ""
    nominal_min: float = 0
    nominal_max: float = 100
    firmware_name: str = ""
    schematic_name: str = ""
    warnings: list[str] = field(default_factory=list)


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    for idx, row in enumerate(rows):
        mapping: dict[str, int] = {}
        for col, cell in enumerate(row):
            key = _norm_header(cell)
            if not key:
                continue
            if key in ("firmware name", "signal name [firmware name]"):
                mapping["firmware_name"] = col
            elif key == "schematic name":
                mapping["schematic_name"] = col
            elif key in ("data type", "signal type"):
                mapping["data_type"] = col
            elif key == "can message id":
                mapping["can_id"] = col
            elif key in ("bit offset", "bit offset "):
                mapping["bit_offset"] = col
            elif key == "units":
                mapping["units"] = col
        if "data_type" in mapping and ("firmware_name" in mapping or "schematic_name" in mapping):
            return idx, mapping
    raise ValueError("Could not locate header row")


def _normalize_type(raw: str) -> str:
    value = raw.strip().lower()
    mapping = {
        "boolean": "bool",
        "bool": "bool",
        "float": "float",
        "integer": "uint8",
        "int": "int32",
        "int32": "int32",
        "uint8": "uint8",
        "uint16": "uint16",
        "uint32": "uint32",
    }
    return mapping.get(value, value)


def _type_bytes(data_type: str) -> int:
    return {
        "bool": 1,
        "uint8": 1,
        "uint16": 2,
        "uint32": 4,
        "int32": 4,
        "float": 4,
    }.get(data_type, 4)


def _normalize_can_id(raw: Any) -> str:
    if raw is None or raw == "":
        return "FFF"
    text = str(raw).strip()
    if text.upper() in ("FFF", "N/A", "NA", "NONE"):
        return "FFF"
    text = text.lower().replace("0x", "").replace("x", "")
    if "." in text:
        text = text.split(".", 1)[0]
    try:
        # CAN IDs are hexadecimal; plain digits like 209 mean 0x209, not decimal 209.
        value = int(text, 16)
        return format(value, "X")
    except ValueError:
        return "FFF"


def _normalize_bit_offset(raw: Any) -> int:
    if raw is None or raw == "":
        return 0
    try:
        return int(float(raw))
    except (TypeError, ValueError):
        return 0


def _should_skip_sheet(name: str, config: dict[str, Any]) -> bool:
    if name in config.get("skip_sheets", []):
        return True
    for pattern in config.get("skip_sheet_patterns", []):
        if pattern in name:
            return True
    return False


def _is_example_row(firmware_name: str, schematic_name: str) -> bool:
    combined = f"{firmware_name} {schematic_name}".lower()
    return combined.startswith("ex") or "example" in combined


def _resolve_software_category(first_col: str, current: str | None, config: dict[str, Any]) -> str | None:
    sections = config.get("software_sections", {})
    for label, category in sections.items():
        if first_col.strip().lower() == label.strip().lower():
            return category
    return current


def _mcc_key(
    firmware_name: str,
    schematic_name: str,
    row_label: str,
    config: dict[str, Any],
) -> str | None:
    title_key = config.get("mcc_title_to_key", {}).get(firmware_name.strip())
    if title_key:
        return title_key
    title_key = config.get("mcc_title_to_key", {}).get(row_label.strip())
    if title_key:
        return title_key
    schematic_key = config.get("mcc_schematic_to_key", {}).get(schematic_name.upper())
    if schematic_key:
        return schematic_key
    return None


def _mppt_key(row_label: str, config: dict[str, Any]) -> str | None:
    labels = config.get("mppt_label_to_key", {})
    label = row_label.strip()
    if label in labels:
        return labels[label]
    for candidate, key in labels.items():
        if candidate.strip() == label:
            return key
    return None


def _resolve_key(
    sheet: str,
    firmware_name: str,
    schematic_name: str,
    row_label: str,
    config: dict[str, Any],
) -> str | None:
    overrides = config.get("key_overrides", {}).get(sheet, {})

    if sheet == "MCC":
        return _mcc_key(firmware_name, schematic_name, row_label, config)

    if sheet == "MPPT":
        return _mppt_key(row_label, config)

    if firmware_name:
        base = firmware_name.strip()
        if base in overrides:
            return overrides[base]
        upper = base.upper()
        if upper in overrides:
            return overrides[upper]
        return base

    if schematic_name:
        return schematic_name.strip().lower()
    return None


def parse_workbook(path: str, config: dict[str, Any]) -> tuple[list[ParsedSignal], list[str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    signals: list[ParsedSignal] = []
    notes: list[str] = []

    for sheet_name in wb.sheetnames:
        if _should_skip_sheet(sheet_name, config):
            notes.append(f"Skipped sheet: {sheet_name}")
            continue
        if sheet_name not in config.get("parse_sheets", []):
            notes.append(f"Ignored sheet (not in parse_sheets): {sheet_name}")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_idx, columns = _find_header_row(rows[:10])
        software_category: str | None = None

        for row_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
            firmware_name = _cell_str(row[columns["firmware_name"]]) if "firmware_name" in columns else ""
            schematic_name = _cell_str(row[columns["schematic_name"]]) if "schematic_name" in columns else ""
            row_label = _cell_str(row[0])

            if sheet_name == "Software":
                maybe_section = _resolve_software_category(row_label, software_category, config)
                if maybe_section and maybe_section != software_category and not firmware_name:
                    software_category = maybe_section
                    continue

            if sheet_name == "Race Strategy" and not firmware_name and not row_label:
                continue

            firmware_name = firmware_name.strip()
            key = _resolve_key(sheet_name, firmware_name, schematic_name, row_label, config)
            if not key:
                if row_label or firmware_name or schematic_name:
                    notes.append(f"{sheet_name} row {row_idx}: no key mapping ({row_label or firmware_name or schematic_name})")
                continue

            known_keys = set(config.get("category_by_key", {}))
            if known_keys and key not in known_keys:
                continue

            if _is_example_row(firmware_name, schematic_name):
                continue

            raw_type = _cell_str(row[columns["data_type"]]) if "data_type" in columns else ""
            if not raw_type:
                notes.append(f"{sheet_name} row {row_idx} ({key}): missing data type")
                continue

            data_type = _normalize_type(raw_type)
            can_id = _normalize_can_id(row[columns["can_id"]]) if "can_id" in columns else "FFF"
            bit_offset = _normalize_bit_offset(row[columns["bit_offset"]]) if "bit_offset" in columns else 0

            if sheet_name in ("Software", "Race Strategy"):
                can_id = "FFF"
                bit_offset = 0

            if sheet_name == "MCC":
                can_id = "FFF"
                bit_offset = 0

            category = config.get("category_by_key", {}).get(key, "")
            if sheet_name == "Software" and software_category:
                category = software_category
            if sheet_name == "Race Strategy":
                category = config.get("race_strategy_category", "Race Strategy;Model Outputs")

            defaults = config.get("signal_defaults", {}).get(key, {})
            units = defaults.get("units", "")
            if "units" in columns:
                sheet_units = _cell_str(row[columns["units"]])
                if sheet_units:
                    units = sheet_units

            warnings: list[str] = []
            if data_type not in {"bool", "uint8", "uint16", "uint32", "int32", "float"}:
                warnings.append(f"unknown data type '{raw_type}'")

            signals.append(
                ParsedSignal(
                    key=key,
                    sheet=sheet_name,
                    row=row_idx,
                    data_type=data_type,
                    can_id=can_id,
                    bit_offset=bit_offset,
                    category=category,
                    units=units,
                    nominal_min=float(defaults.get("min", 0)),
                    nominal_max=float(defaults.get("max", 100)),
                    firmware_name=firmware_name,
                    schematic_name=schematic_name,
                    warnings=warnings,
                )
            )

    wb.close()
    return signals, notes
